from dataclasses import dataclass
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
import json
import os
from pathlib import Path
import shutil
import sys
import threading
import time
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from urllib.parse import parse_qs, urlparse

from openpyxl import load_workbook

from fill_excel import fill_workbook, load_data, safe_filename


APP_NAME = "Export Genius 자동 저장 도구"
APP_VERSION = "0.3.50"
EXPECTED_EXTENSION_VERSION = "0.3.50"
LOCAL_API_VERSION = 1


def resource_dir() -> Path:
    bundled = getattr(sys, "_MEIPASS", None)
    if bundled:
        return Path(bundled)
    return Path(__file__).resolve().parent


RESOURCE_DIR = resource_dir()
IS_FROZEN = bool(getattr(sys, "frozen", False))
ROAMING_APP_DIR = Path(os.environ.get("APPDATA", Path.home() / "AppData" / "Roaming")) / "ExportGenius"
LOCAL_APP_DIR = Path(os.environ.get("LOCALAPPDATA", Path.home() / "AppData" / "Local")) / "ExportGenius"
SETTINGS_PATH = ROAMING_APP_DIR / "settings.json"
LOG_DIR = LOCAL_APP_DIR / "logs"
APP_LOG_PATH = LOG_DIR / "app.log"
RAW_LOG_PATH = LOG_DIR / "extension-raw.jsonl"
DEFAULT_DOWNLOADS = Path.home() / "Downloads"
DEFAULT_TEMPLATE = RESOURCE_DIR / "template.xlsx"
DEFAULT_OUTPUT_DIR = (
    Path.home() / "Documents" / "Export Genius Output"
    if IS_FROZEN
    else RESOURCE_DIR / "output"
)
LOCAL_SERVER_HOST = "127.0.0.1"
LOCAL_SERVER_PORT = 8765


def load_user_settings() -> dict:
    try:
        data = json.loads(SETTINGS_PATH.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except (OSError, json.JSONDecodeError):
        return {}


def save_user_settings(settings: dict) -> None:
    ROAMING_APP_DIR.mkdir(parents=True, exist_ok=True)
    temporary_path = SETTINGS_PATH.with_suffix(".tmp")
    temporary_path.write_text(json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary_path.replace(SETTINGS_PATH)


def rotate_log_file(path: Path, max_bytes: int = 5 * 1024 * 1024) -> None:
    try:
        if not path.exists() or path.stat().st_size < max_bytes:
            return
        backup = path.with_suffix(path.suffix + ".1")
        if backup.exists():
            backup.unlink()
        path.replace(backup)
    except OSError:
        return


@dataclass(frozen=True)
class ExportTask:
    company: str
    hscode: str
    min_usd: str
    max_usd: str
    company_order: int = 0
    hscode_order: int = 0


def latest_downloaded_xlsx() -> Path | None:
    files = sorted(DEFAULT_DOWNLOADS.glob("*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)
    return files[0] if files else None


def normalize_hscode(value) -> str:
    if value in (None, ""):
        return ""

    text = str(value).strip()
    if not text:
        return ""
    if text.endswith(".0"):
        text = text[:-2]

    return text.zfill(6) if text.isdigit() and len(text) < 6 else text


def header_index(headers: list[str], names: list[str]) -> int | None:
    normalized = {str(header).strip().lower(): index for index, header in enumerate(headers)}

    for name in names:
        index = normalized.get(name.strip().lower())
        if index is not None:
            return index

    return None


def read_input_tasks(input_path: Path) -> list[ExportTask]:
    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return []

    headers = [str(value or "").strip() for value in rows[0]]
    company_col = header_index(
        headers,
        [
            "\uae30\uc5c5\uba85",  # company name
            "\ud68c\uc0ac\uba85",
            "\uc5c5\uccb4\uba85",
            "company",
            "company name",
        ],
    )
    min_col = header_index(headers, ["USD min", "min", "minimum"])
    max_col = header_index(headers, ["USD max", "max", "maximum"])
    hscode_cols = [
        index
        for index, header in enumerate(headers)
        if str(header).strip().lower().startswith(("hsk", "hs code", "hscode", "hs\ucf54\ub4dc"))
    ]

    if company_col is None:
        raise ValueError("업체명 열을 찾지 못했습니다. 헤더에 '기업명', '회사명', '업체명' 중 하나가 필요합니다.")
    if not hscode_cols:
        raise ValueError("HS코드 열을 찾지 못했습니다. 헤더에 HSK1, HSK2, HSK3 같은 열이 필요합니다.")

    tasks: list[ExportTask] = []

    company_orders: dict[str, int] = {}
    hscode_orders: dict[str, int] = {}

    for row in rows[1:]:
        company = str(row[company_col] or "").strip()
        if not company:
            continue

        if company not in company_orders:
            company_orders[company] = len(company_orders) + 1

        min_usd = str(row[min_col] or "50000").strip() if min_col is not None else "50000"
        max_usd = str(row[max_col] or "5000000").strip() if max_col is not None else "5000000"

        for hscode_col in hscode_cols:
            hscode = normalize_hscode(row[hscode_col] if hscode_col < len(row) else "")
            if hscode:
                hscode_orders[company] = hscode_orders.get(company, 0) + 1
                tasks.append(
                    ExportTask(
                        company=company,
                        hscode=hscode,
                        min_usd=min_usd,
                        max_usd=max_usd,
                        company_order=company_orders[company],
                        hscode_order=hscode_orders[company],
                    )
                )

    return tasks


def task_output_dir(output_root: Path, task: ExportTask) -> Path:
    company_name = safe_filename(task.company)
    hscode_name = safe_filename(task.hscode)

    if task.company_order > 0:
        company_name = f"{task.company_order:03d}. {company_name}"
    if task.hscode_order > 0:
        hscode_name = f"{task.hscode_order:03d}. {hscode_name}"

    return output_root / company_name / hscode_name


def legacy_task_output_dirs(output_root: Path, task: ExportTask) -> list[Path]:
    return [
        output_root / safe_filename(task.company) / safe_filename(task.hscode),
        output_root / f"{task.company_order:03d}. {safe_filename(task.company)}" / safe_filename(task.hscode),
        output_root / safe_filename(task.company) / f"{task.hscode_order:03d}. {safe_filename(task.hscode)}",
    ]


def task_output_dirs(output_root: Path, task: ExportTask) -> list[Path]:
    primary = task_output_dir(output_root, task)
    paths = [primary]
    for path in legacy_task_output_dirs(output_root, task):
        if path not in paths:
            paths.append(path)
    return paths


def count_xlsx_files(folder: Path) -> int:
    return len(list(folder.glob("*.xlsx"))) if folder.exists() else 0


def count_task_xlsx_files(output_root: Path, task: ExportTask) -> int:
    seen: set[Path] = set()
    for folder in task_output_dirs(output_root, task):
        if not folder.exists():
            continue
        for path in folder.glob("*.xlsx"):
            seen.add(path.resolve())
    return len(seen)


def buyer_name_from_xlsx_path(path: Path) -> str:
    stem = strip_download_suffix(path.stem).strip()
    prefix, separator, remainder = stem.partition(".")
    if separator and prefix.strip().isdigit() and remainder.strip():
        return remainder.strip()

    return stem


def xlsx_index_from_path(path: Path) -> int:
    prefix = path.stem.split(".", 1)[0].strip()
    return int(prefix) if prefix.isdigit() else 0


def last_saved_buyer_for_task(output_root: Path, task: ExportTask) -> dict:
    candidates: list[tuple[int, str, Path]] = []
    seen_paths: set[Path] = set()

    for folder in task_output_dirs(output_root, task):
        if not folder.exists():
            continue

        for path in folder.glob("*.xlsx"):
            resolved = path.resolve()
            if resolved in seen_paths:
                continue

            seen_paths.add(resolved)
            buyer_name = buyer_name_from_xlsx_path(path)
            if buyer_name:
                candidates.append((xlsx_index_from_path(path), buyer_name, path))

    if not candidates:
        return {}

    index, buyer_name, path = max(candidates, key=lambda item: (item[0], item[2].stat().st_mtime))
    return {
        "index": str(index),
        "buyerName": buyer_name,
        "fileName": path.name,
    }


def saved_buyer_names_for_task(output_root: Path, task: ExportTask) -> list[str]:
    buyers: list[str] = []
    seen_paths: set[Path] = set()
    seen_keys: set[str] = set()

    for folder in task_output_dirs(output_root, task):
        if not folder.exists():
            continue

        for path in sorted(folder.glob("*.xlsx")):
            resolved = path.resolve()
            if resolved in seen_paths:
                continue

            seen_paths.add(resolved)
            buyer_name = buyer_name_from_xlsx_path(path)
            buyer_key = normalize_name_key(buyer_name)
            if not buyer_name or not buyer_key or buyer_key in seen_keys:
                continue

            seen_keys.add(buyer_key)
            buyers.append(buyer_name)

    return buyers


def wait_until_json_ready(path: Path, attempts: int = 8, delay_seconds: float = 0.5) -> dict:
    previous_size = -1

    for _ in range(attempts):
        if not path.exists() or path.name.endswith(".crdownload"):
            time.sleep(delay_seconds)
            continue

        size = path.stat().st_size
        if size == previous_size and size > 0:
            return load_data(path)

        previous_size = size
        time.sleep(delay_seconds)

    return load_data(path)


def next_output_path(folder: Path, buyer_name: str, related_folders: list[Path] | None = None) -> Path:
    existing_indexes: list[int] = []
    existing_paths: set[Path] = set()
    folders = related_folders or [folder]

    for related_folder in folders:
        if not related_folder.exists():
            continue

        for path in related_folder.glob("*.xlsx"):
            resolved = path.resolve()
            if resolved in existing_paths:
                continue

            existing_paths.add(resolved)
            prefix = path.stem.split(".", 1)[0].strip()
            if prefix.isdigit():
                existing_indexes.append(int(prefix))

    index = max(existing_indexes + [len(existing_paths)], default=0) + 1
    buyer_file_name = safe_filename(buyer_name)
    output_path = folder / f"{index}. {buyer_file_name}.xlsx"

    while output_path.exists():
        index += 1
        output_path = folder / f"{index}. {buyer_file_name}.xlsx"

    return output_path


def normalize_name_key(value: str) -> str:
    return "".join(ch.lower() for ch in str(value or "") if ch.isalnum())


def strip_download_suffix(stem: str) -> str:
    text = str(stem or "").strip()
    if text.endswith(")") and " (" in text:
        base, suffix = text.rsplit(" (", 1)
        if suffix[:-1].isdigit():
            return base

    return text


def json_filename_matches_buyer(json_path: Path, buyer_name: str) -> bool:
    file_key = normalize_name_key(strip_download_suffix(json_path.stem))
    buyer_key = normalize_name_key(safe_filename(buyer_name))

    if not file_key or not buyer_key:
        return False

    return file_key == buyer_key or file_key.startswith(buyer_key[:80]) or buyer_key.startswith(file_key[:80])


def move_json_to_folder(json_path: Path, folder: Path) -> Path:
    folder.mkdir(parents=True, exist_ok=True)
    target = folder / json_path.name

    if target.exists():
        stem = target.stem
        suffix = target.suffix
        counter = 2
        while target.exists():
            target = folder / f"{stem} ({counter}){suffix}"
            counter += 1

    shutil.move(str(json_path), str(target))
    return target


def matching_json_candidates(folder: Path, expected_filename: str) -> list[Path]:
    expected = Path(expected_filename)
    expected_stem = strip_download_suffix(expected.stem)
    expected_key = normalize_name_key(expected_stem)

    candidates: list[Path] = []
    for path in folder.glob("*.json"):
        if not path.is_file():
            continue

        if path.name == expected_filename:
            candidates.append(path)
            continue

        if normalize_name_key(strip_download_suffix(path.stem)) == expected_key:
            candidates.append(path)

    return sorted(candidates, key=lambda item: item.stat().st_mtime, reverse=True)


class LocalApiHandler(BaseHTTPRequestHandler):
    app: "ExcelGui"

    def log_message(self, format: str, *args) -> None:
        return

    def send_json(self, status: int, payload: dict) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()
        self.wfile.write(body)

    def do_OPTIONS(self) -> None:
        self.send_json(200, {"ok": True})

    def do_GET(self) -> None:
        parsed = urlparse(self.path)

        if parsed.path == "/health":
            self.send_json(200, self.app.local_health_payload())
            return

        if parsed.path == "/task":
            self.send_json(200, self.app.local_task_payload())
            return

        if parsed.path == "/queue/current":
            self.send_json(200, self.app.local_queue_current_payload())
            return

        if parsed.path == "/queue/complete":
            self.send_json(200, self.app.local_queue_complete_payload(parse_qs(parsed.query)))
            return

        if parsed.path == "/queue/fail":
            self.send_json(200, self.app.local_queue_fail_payload(parse_qs(parsed.query)))
            return

        if parsed.path == "/log":
            self.send_json(200, self.app.local_log_payload(parse_qs(parsed.query)))
            return

        if parsed.path == "/command":
            self.send_json(200, self.app.local_command_payload())
            return

        if parsed.path == "/command-result":
            self.send_json(200, self.app.local_command_result_payload(parse_qs(parsed.query)))
            return

        if parsed.path == "/stop-request":
            self.send_json(200, self.app.local_stop_request_payload())
            return

        if parsed.path == "/download-confirm":
            self.send_json(200, self.app.local_download_confirm_payload(parse_qs(parsed.query)))
            return

        self.send_json(404, {"ok": False, "reason": "Unknown endpoint."})

    def do_POST(self) -> None:
        parsed = urlparse(self.path)

        if parsed.path == "/raw-log":
            length = int(self.headers.get("Content-Length") or "0")
            body = self.rfile.read(length).decode("utf-8", errors="replace") if length > 0 else ""
            self.send_json(200, self.app.local_raw_log_payload(body))
            return

        self.send_json(404, {"ok": False, "reason": "Unknown endpoint."})


class ExcelGui:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title(f"{APP_NAME} {APP_VERSION}")
        self.root.geometry("980x620")
        self.root.minsize(860, 520)

        self.settings = load_user_settings()
        default_input = latest_downloaded_xlsx()
        configured_input = str(self.settings.get("inputFile") or "")
        configured_template = str(self.settings.get("templateFile") or "")
        configured_downloads = str(self.settings.get("downloadDir") or "")
        configured_output = str(self.settings.get("outputDir") or "")
        self.input_var = tk.StringVar(value=configured_input or (str(default_input) if default_input else ""))
        self.template_var = tk.StringVar(
            value=configured_template or (str(DEFAULT_TEMPLATE) if DEFAULT_TEMPLATE.exists() else "")
        )
        self.download_dir_var = tk.StringVar(value=configured_downloads or str(DEFAULT_DOWNLOADS))
        self.output_dir_var = tk.StringVar(value=configured_output or str(DEFAULT_OUTPUT_DIR))
        self.task_index_var = tk.StringVar(value="1")
        self.target_count_var = tk.StringVar(value="60")
        self.default_target_var = tk.StringVar(value=str(self.settings.get("defaultTarget") or "60"))
        self.watch_timeout_var = tk.StringVar(value=str(self.settings.get("watchTimeout") or "300"))
        self.single_json_var = tk.StringVar()
        self.status_var = tk.StringVar(value="1단계: 파일과 저장 위치를 선택하세요.")
        self.extension_status_var = tk.StringVar(value="확장프로그램 연결 대기")

        self.tasks: list[ExportTask] = []
        self.task_rows: list[dict] = []
        self.queue_indices: list[int] = []
        self.queue_position = 0
        self.queue_active = False
        self.queue_run_id = ""
        self.queue_task_ids: dict[int, str] = {}
        self.queue_lock = threading.Lock()
        self.watch_thread: threading.Thread | None = None
        self.stop_event = threading.Event()
        self.local_server: ThreadingHTTPServer | None = None
        self.current_task_payload: dict | None = None
        self.payload_lock = threading.Lock()
        self.pending_command: dict | None = None
        self.command_counter = 0
        self.command_lock = threading.Lock()
        self.stop_requested = False
        self.raw_logs: list[dict] = []
        self.raw_log_lock = threading.Lock()
        self.file_log_lock = threading.Lock()
        self.last_extension_version = ""
        self.last_extension_seen_at = ""
        self.server_start_error = ""

        try:
            LOG_DIR.mkdir(parents=True, exist_ok=True)
            rotate_log_file(APP_LOG_PATH)
            rotate_log_file(RAW_LOG_PATH)
        except OSError:
            pass

        self.build()
        if not self.start_local_server():
            self.root.after(100, self.show_server_start_error)
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def build(self) -> None:
        container = ttk.Frame(self.root, padding=12)
        container.pack(fill="both", expand=True)
        container.rowconfigure(1, weight=1)
        container.columnconfigure(0, weight=1)

        title = ttk.Label(container, text="Export Genius 자동 저장 도구", font=("Segoe UI", 14, "bold"))
        title.grid(row=0, column=0, sticky="w")

        self.notebook = ttk.Notebook(container)
        self.notebook.grid(row=1, column=0, sticky="nsew", pady=(10, 8))

        self.setup_tab = ttk.Frame(self.notebook, padding=12)
        self.tasks_tab = ttk.Frame(self.notebook, padding=12)
        self.watch_tab = ttk.Frame(self.notebook, padding=12)
        self.tools_tab = ttk.Frame(self.notebook, padding=12)
        self.help_tab = ttk.Frame(self.notebook, padding=12)

        self.notebook.add(self.setup_tab, text="1 파일 선택")
        self.notebook.add(self.tasks_tab, text="2 작업 확인")
        self.notebook.add(self.watch_tab, text="3 모니터링")
        self.notebook.add(self.tools_tab, text="4 보조 기능")
        self.notebook.add(self.help_tab, text="5 사용 안내")

        self.build_setup_tab()
        self.build_tasks_tab()
        self.build_watch_tab()
        self.build_tools_tab()
        self.build_help_tab()

        status = ttk.Label(container, textvariable=self.status_var, anchor="w")
        status.grid(row=2, column=0, sticky="ew")

    def start_local_server(self) -> bool:
        try:
            handler = type("ExportGeniusLocalApiHandler", (LocalApiHandler,), {"app": self})
            self.local_server = ThreadingHTTPServer((LOCAL_SERVER_HOST, LOCAL_SERVER_PORT), handler)
            self.local_server.daemon_threads = True
            thread = threading.Thread(target=self.local_server.serve_forever, daemon=True)
            thread.start()
            return True
        except OSError as error:
            self.local_server = None
            self.server_start_error = (
                f"로컬 연결 서버를 시작하지 못했습니다.\n\n"
                f"주소: http://{LOCAL_SERVER_HOST}:{LOCAL_SERVER_PORT}\n"
                f"원인: {error}\n\n"
                "프로그램이 이미 실행 중인지 확인한 뒤 다시 실행하세요."
            )
            self.status_var.set("프로그램 시작 실패: 로컬 연결 포트 사용 중")
            return False

    def show_server_start_error(self) -> None:
        if not self.server_start_error:
            return
        messagebox.showerror("프로그램을 시작할 수 없음", self.server_start_error)
        self.root.destroy()

    def current_settings(self) -> dict:
        return {
            "inputFile": self.input_var.get().strip(),
            "templateFile": self.template_var.get().strip(),
            "downloadDir": self.download_dir_var.get().strip(),
            "outputDir": self.output_dir_var.get().strip(),
            "defaultTarget": self.default_target_var.get().strip(),
            "watchTimeout": self.watch_timeout_var.get().strip(),
        }

    def persist_settings(self) -> None:
        try:
            save_user_settings(self.current_settings())
        except OSError as error:
            self.write_file_log(f"설정 저장 실패: {error}")

    def on_close(self) -> None:
        self.stop_event.set()
        self.persist_settings()
        if self.local_server:
            self.local_server.shutdown()
            self.local_server.server_close()
        self.root.destroy()

    def local_health_payload(self) -> dict:
        return {
            "ok": True,
            "app": APP_NAME,
            "version": APP_VERSION,
            "apiVersion": LOCAL_API_VERSION,
            "expectedExtensionVersion": EXPECTED_EXTENSION_VERSION,
            "server": f"http://{LOCAL_SERVER_HOST}:{LOCAL_SERVER_PORT}",
        }

    def local_task_payload(self) -> dict:
        with self.payload_lock:
            task = dict(self.current_task_payload or {})

        if not task:
            return {
                "ok": False,
                "reason": "현재 선택된 작업이 없습니다. GUI에서 작업을 선택한 뒤 다시 확인하세요.",
            }

        return {
            "ok": True,
            "task": task,
        }

    def task_payload(
        self,
        task: ExportTask,
        task_number: int,
        target_count: int | str,
        queue_task_id: str | None = None,
    ) -> dict:
        output_root = Path(self.output_dir_var.get())
        already_saved = count_task_xlsx_files(output_root, task)
        already_saved_buyers = saved_buyer_names_for_task(output_root, task)
        last_saved_buyer = last_saved_buyer_for_task(output_root, task)
        remaining_count = max(0, int(target_count) - already_saved)

        return {
            "taskNumber": task_number,
            "company": task.company,
            "hsCode": task.hscode,
            "minValue": task.min_usd,
            "maxValue": task.max_usd,
            "targetCount": str(target_count),
            "alreadySaved": str(already_saved),
            "remainingCount": str(remaining_count),
            "alreadySavedBuyers": already_saved_buyers,
            "lastSavedBuyer": last_saved_buyer,
            "queueTaskId": queue_task_id or "",
        }

    def set_current_task_payload(self, payload: dict | None) -> None:
        with self.payload_lock:
            self.current_task_payload = dict(payload) if payload else None

    def local_queue_current_payload(self) -> dict:
        with self.queue_lock:
            if not self.queue_active:
                return {"ok": True, "done": True, "reason": "No active queue."}

            if self.queue_position >= len(self.queue_indices):
                self.queue_active = False
                self.queue_task_ids = {}
                self.set_current_task_payload(None)
                return {"ok": True, "done": True, "reason": "Queue completed."}

            row_index = self.queue_indices[self.queue_position]
            row = self.task_rows[row_index]
            task = row["task"]
            target = row["target"]
            row["status"] = "진행중"
            queue_task_id = self.queue_task_ids.get(row_index, "")
            payload = self.task_payload(task, row_index + 1, target, queue_task_id)
            self.set_current_task_payload(payload)

            self.root.after(0, self.refresh_task_tree)
            return {
                "ok": True,
                "done": False,
                "queuePosition": self.queue_position + 1,
                "queueTotal": len(self.queue_indices),
                "task": payload,
            }

    def local_queue_complete_payload(self, query: dict[str, list[str]]) -> dict:
        with self.queue_lock:
            if not self.queue_active or self.queue_position >= len(self.queue_indices):
                return {"ok": False, "reason": "완료 처리할 자동 작업이 없습니다."}

            row_index = self.queue_indices[self.queue_position]
            row = self.task_rows[row_index]
            task = row["task"]
            target = int(row["target"])
            expected_task_id = self.queue_task_ids.get(row_index, "")

            requested_task_id = (query.get("taskId") or [""])[0]
            if expected_task_id and requested_task_id != expected_task_id:
                reason = "완료 보고가 현재 작업과 일치하지 않아 무시했습니다."
                self.post_watch_message(
                    f"자동 작업 완료 보고 무시: {task.company} / {task.hscode} - {reason}",
                    f"오류: {reason}",
                )
                return {
                    "ok": False,
                    "stale": True,
                    "reason": reason,
                    "expectedTaskId": expected_task_id,
                    "receivedTaskId": requested_task_id,
                }

        task_dir = task_output_dir(Path(self.output_dir_var.get()), task)
        deadline = time.time() + 20
        saved_count = count_task_xlsx_files(Path(self.output_dir_var.get()), task)
        while saved_count < target and time.time() < deadline:
            time.sleep(0.5)
            saved_count = count_task_xlsx_files(Path(self.output_dir_var.get()), task)

        if saved_count < target:
            reason = f"엑셀 파일이 목표 수량보다 적습니다. 현재 {saved_count}/{target}개"
            with self.queue_lock:
                self.task_rows[row_index]["status"] = "오류"
                self.queue_active = False
            self.root.after(0, self.refresh_task_tree)
            self.post_watch_message(f"자동 작업 중단: {task.company} / {task.hscode} - {reason}", f"오류: {reason}")
            return {"ok": False, "reason": reason, "savedCount": saved_count, "targetCount": target}

        with self.queue_lock:
            self.task_rows[row_index]["status"] = f"완료 {saved_count}/{target}"
            self.queue_position += 1
            done = self.queue_position >= len(self.queue_indices)

            if done:
                self.queue_active = False
                self.queue_task_ids = {}
                next_payload = None
            else:
                next_row_index = self.queue_indices[self.queue_position]
                next_row = self.task_rows[next_row_index]
                next_row["status"] = "준비됨"
                next_payload = self.task_payload(
                    next_row["task"],
                    next_row_index + 1,
                    next_row["target"],
                    self.queue_task_ids.get(next_row_index, ""),
                )

            self.set_current_task_payload(next_payload)

        self.root.after(0, self.refresh_task_tree)
        self.post_watch_message(f"자동 작업 완료: {task.company} / {task.hscode} - 엑셀 {saved_count}개")

        return {
            "ok": True,
            "done": done,
            "savedCount": saved_count,
            "targetCount": target,
            "nextTask": next_payload,
        }

    def local_queue_fail_payload(self, query: dict[str, list[str]]) -> dict:
        reason = (query.get("reason") or [""])[0] or "확장프로그램 작업 실패"
        requested_task_id = (query.get("taskId") or [""])[0]

        with self.queue_lock:
            if self.queue_active and self.queue_position < len(self.queue_indices):
                row_index = self.queue_indices[self.queue_position]
                expected_task_id = self.queue_task_ids.get(row_index, "")
                if expected_task_id and requested_task_id != expected_task_id:
                    stale_reason = "실패 보고가 현재 작업과 일치하지 않아 무시했습니다."
                    self.post_watch_message(f"자동 작업 실패 보고 무시: {stale_reason}", f"오류: {stale_reason}")
                    return {
                        "ok": False,
                        "stale": True,
                        "reason": stale_reason,
                        "expectedTaskId": expected_task_id,
                        "receivedTaskId": requested_task_id,
                    }
                self.task_rows[row_index]["status"] = "오류"
            self.queue_active = False
            self.queue_task_ids = {}
            self.set_current_task_payload(None)

        self.root.after(0, self.refresh_task_tree)
        self.post_watch_message(f"자동 작업 중단: {reason}", f"오류: {reason}")
        return {"ok": True, "stopped": True, "reason": reason}

    def local_log_payload(self, query: dict[str, list[str]]) -> dict:
        message = (query.get("message") or [""])[0].strip()
        status = (query.get("status") or [""])[0].strip() or None

        if message:
            self.post_watch_message(message, status)

        return {"ok": True}

    def local_command_payload(self) -> dict:
        with self.command_lock:
            if not self.pending_command or self.pending_command.get("claimed"):
                return {"ok": True, "command": None}

            self.pending_command["claimed"] = True
            command = dict(self.pending_command)

        self.post_watch_message(f"확장프로그램이 명령을 수신했습니다: {command.get('action')}")
        return {"ok": True, "command": command}

    def local_command_result_payload(self, query: dict[str, list[str]]) -> dict:
        command_id = (query.get("id") or [""])[0]
        ok = (query.get("ok") or [""])[0].lower() == "true"
        message = (query.get("message") or [""])[0].strip()

        with self.command_lock:
            if self.pending_command and str(self.pending_command.get("id")) == str(command_id):
                self.pending_command = None

        if message:
            self.post_watch_message(message, "완료" if ok else "오류 발생")

        return {"ok": True}

    def local_raw_log_payload(self, body: str) -> dict:
        try:
            payload = json.loads(body) if body else {}
        except json.JSONDecodeError:
            payload = {"raw": body}

        entry = {
            "receivedAt": time.strftime("%Y-%m-%d %H:%M:%S"),
            "payload": payload,
        }

        with self.raw_log_lock:
            self.raw_logs.append(entry)
            self.raw_logs = self.raw_logs[-30:]
            count = len(self.raw_logs)

        self.write_raw_log(entry)

        if payload.get("type") == "extension-ready":
            version = str(payload.get("extensionVersion") or "").strip()
            self.last_extension_version = version
            self.last_extension_seen_at = entry["receivedAt"]
            if version == EXPECTED_EXTENSION_VERSION:
                status = f"확장프로그램 연결됨 ({version})"
            else:
                status = f"확장프로그램 버전 불일치 ({version or '확인 불가'})"
            self.root.after(0, lambda value=status: self.extension_status_var.set(value))

        return {"ok": True, "count": count}

    def local_stop_request_payload(self) -> dict:
        return {"ok": True, "stopRequested": bool(self.stop_requested)}

    def local_download_confirm_payload(self, query: dict[str, list[str]]) -> dict:
        filename = (query.get("filename") or [""])[0]
        buyer_name = (query.get("buyer") or [""])[0]
        hs_code = normalize_hscode((query.get("hsCode") or [""])[0])
        require_excel = (query.get("requireExcel") or [""])[0] == "1"
        timeout_seconds = int((query.get("timeout") or ["20"])[0] or "20")
        download_dir = Path(self.download_dir_var.get())
        deadline = time.time() + max(1, min(timeout_seconds, 60))
        last_error = ""

        if not filename:
            return {"ok": False, "reason": "확인할 JSON 파일명이 비어 있습니다."}
        if not download_dir.exists():
            return {"ok": False, "reason": f"다운로드 폴더를 찾을 수 없습니다: {download_dir}"}

        while time.time() < deadline:
            for json_path in matching_json_candidates(download_dir, filename):
                try:
                    data = wait_until_json_ready(json_path, attempts=4, delay_seconds=0.25)
                    json_buyer_name = data.get("excel", {}).get("pink", {}).get("Company_Name") or json_path.stem
                    json_hscode = normalize_hscode(data.get("hsCode"))

                    if buyer_name and not json_filename_matches_buyer(json_path, buyer_name):
                        last_error = f"파일명과 예상 바이어명({buyer_name})이 일치하지 않습니다."
                        continue
                    if not json_filename_matches_buyer(json_path, json_buyer_name):
                        last_error = f"파일명과 JSON 내부 바이어명({json_buyer_name})이 일치하지 않습니다."
                        continue
                    if hs_code and json_hscode and json_hscode != hs_code:
                        last_error = f"JSON HS코드({json_hscode})가 예상 HS코드({hs_code})와 다릅니다."
                        continue

                    conversion = self.convert_queue_json_if_needed(json_path, hs_code)
                    if not conversion["ok"]:
                        return conversion
                    if require_excel and not conversion.get("converted"):
                        return {
                            "ok": False,
                            "reason": "JSON은 확인했지만 현재 자동 작업과 연결되지 않아 엑셀로 변환하지 못했습니다.",
                            "filename": json_path.name,
                        }
                    if require_excel and not Path(str(conversion.get("outputFile") or "")).exists():
                        return {
                            "ok": False,
                            "reason": "엑셀 파일 생성 완료를 확인하지 못했습니다.",
                            "filename": json_path.name,
                            "outputFile": conversion.get("outputFile"),
                        }

                    return {
                        "ok": True,
                        "filename": json_path.name,
                        "path": str(conversion.get("jsonPath") or json_path),
                        "buyer": json_buyer_name,
                        "hsCode": json_hscode,
                        "converted": conversion.get("converted", False),
                        "outputCount": conversion.get("outputCount"),
                        "outputFile": conversion.get("outputFile"),
                    }
                except Exception as error:
                    last_error = str(error)

            time.sleep(0.5)

        return {
            "ok": False,
            "reason": last_error or f"{timeout_seconds}초 안에 다운로드된 JSON 파일을 확인하지 못했습니다.",
            "expectedFilename": filename,
            "downloadDir": str(download_dir),
        }

    def convert_queue_json_if_needed(self, json_path: Path, hs_code: str) -> dict:
        with self.queue_lock:
            if not self.queue_active or self.queue_position >= len(self.queue_indices):
                return {"ok": True, "converted": False, "jsonPath": str(json_path)}

            row_index = self.queue_indices[self.queue_position]
            row = self.task_rows[row_index]
            task = row["task"]
            target = int(row["target"])

        if hs_code and normalize_hscode(task.hscode) != hs_code:
            return {"ok": True, "converted": False, "jsonPath": str(json_path)}

        task_dir = task_output_dir(Path(self.output_dir_var.get()), task)
        try:
            template_path = Path(self.template_var.get())
            if not template_path.exists():
                raise FileNotFoundError("결과 템플릿 파일을 찾을 수 없습니다.")

            task_dir.mkdir(parents=True, exist_ok=True)
            output_path = self.process_downloaded_json(task, template_path, task_dir, json_path)
            output_count = count_xlsx_files(task_dir)

            with self.queue_lock:
                if self.queue_active and self.queue_position < len(self.queue_indices):
                    active_index = self.queue_indices[self.queue_position]
                    if active_index == row_index:
                        self.task_rows[row_index]["status"] = f"저장 {output_count}/{target}"

            self.root.after(0, self.refresh_task_tree)
            self.post_watch_message(f"[엑셀 생성] {output_count}/{target}: {output_path.name}")
            return {
                "ok": True,
                "converted": True,
                "jsonPath": str(task_dir / "_json_backup" / json_path.name),
                "outputFile": str(output_path),
                "outputCount": output_count,
            }
        except Exception as error:
            failed_path = json_path
            if json_path.exists():
                failed_path = move_json_to_folder(json_path, task_dir / "_failed_json")

            with self.queue_lock:
                self.task_rows[row_index]["status"] = "오류"
                self.queue_active = False
                self.set_current_task_payload(None)

            self.root.after(0, self.refresh_task_tree)
            return {
                "ok": False,
                "reason": f"JSON 확인 후 엑셀 변환에 실패했습니다: {error}",
                "failedJson": str(failed_path),
            }

    def update_current_task_payload(self, task: ExportTask, task_number: int) -> None:
        payload = self.task_payload(task, task_number, self.target_count_var.get())

        with self.payload_lock:
            self.current_task_payload = payload

    def build_setup_tab(self) -> None:
        self.setup_tab.columnconfigure(1, weight=1)

        intro = (
            "작업 목록 엑셀, 결과 템플릿, 최종 저장 폴더를 선택하세요. "
            "브라우저 다운로드 폴더는 기본 Downloads 폴더를 자동으로 모니터링합니다."
        )
        ttk.Label(self.setup_tab, text=intro, wraplength=820).grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 14))

        self.add_file_row(self.setup_tab, 1, "작업 목록 엑셀", self.input_var, self.choose_input)
        self.add_file_row(self.setup_tab, 2, "결과 템플릿", self.template_var, self.choose_template)
        self.add_file_row(self.setup_tab, 3, "최종 저장 폴더", self.output_dir_var, self.choose_output_dir)

        actions = ttk.Frame(self.setup_tab)
        actions.grid(row=4, column=0, columnspan=3, sticky="ew", pady=(18, 0))
        ttk.Button(actions, text="선택 내용 확인", command=self.validate_setup).pack(side="left")
        ttk.Button(actions, text="회사/HS코드 폴더 만들기", command=self.create_folders).pack(side="left", padx=(8, 0))
        ttk.Button(actions, text="다음: 작업 확인", command=self.go_tasks).pack(side="left", padx=(8, 0))

        self.setup_summary = tk.Text(self.setup_tab, height=10, wrap="word")
        self.setup_summary.grid(row=5, column=0, columnspan=3, sticky="nsew", pady=(14, 0))
        self.setup_tab.rowconfigure(5, weight=1)

    def build_tasks_tab(self) -> None:
        self.tasks_tab.rowconfigure(1, weight=1)
        self.tasks_tab.columnconfigure(0, weight=1)

        toolbar = ttk.Frame(self.tasks_tab)
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        ttk.Button(toolbar, text="전체 선택", command=self.select_all_tasks).pack(side="left", padx=(8, 0))
        ttk.Button(toolbar, text="전체 해제", command=self.clear_all_tasks).pack(side="left", padx=(8, 0))
        ttk.Label(toolbar, text="기본 목표").pack(side="left", padx=(18, 4))
        ttk.Entry(toolbar, textvariable=self.default_target_var, width=8).pack(side="left")
        ttk.Button(toolbar, text="선택 항목에 목표 적용", command=self.apply_target_to_selected_rows).pack(side="left", padx=(8, 0))
        ttk.Button(toolbar, text="다음: 모니터링", command=self.prepare_selected_queue).pack(side="left", padx=(8, 0))

        columns = ("checked", "number", "company", "hscode", "min_usd", "max_usd", "target", "status")
        self.tasks_tree = ttk.Treeview(self.tasks_tab, columns=columns, show="headings", selectmode="extended")
        headings = {
            "checked": "선택",
            "number": "번호",
            "company": "업체명",
            "hscode": "HS코드",
            "min_usd": "최소 USD",
            "max_usd": "최대 USD",
            "target": "목표",
            "status": "상태",
        }
        widths = {
            "checked": 56,
            "number": 52,
            "company": 260,
            "hscode": 90,
            "min_usd": 90,
            "max_usd": 100,
            "target": 64,
            "status": 120,
        }
        for column in columns:
            self.tasks_tree.heading(column, text=headings[column])
            self.tasks_tree.column(column, width=widths[column], anchor="center" if column != "company" else "w")

        y_scroll = ttk.Scrollbar(self.tasks_tab, orient="vertical", command=self.tasks_tree.yview)
        x_scroll = ttk.Scrollbar(self.tasks_tab, orient="horizontal", command=self.tasks_tree.xview)
        self.tasks_tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        self.tasks_tree.grid(row=1, column=0, sticky="nsew")
        y_scroll.grid(row=1, column=1, sticky="ns")
        x_scroll.grid(row=2, column=0, sticky="ew")
        self.tasks_tree.bind("<ButtonRelease-1>", self.on_task_tree_click)

    def build_watch_tab(self) -> None:
        self.watch_tab.rowconfigure(1, weight=1)
        self.watch_tab.columnconfigure(0, weight=1)

        actions = ttk.Frame(self.watch_tab)
        actions.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        ttk.Button(actions, text="모니터링 시작", command=self.start_gui_automation).pack(side="left")
        ttk.Button(actions, text="모니터링 중단", command=self.stop_gui_automation).pack(side="left", padx=(8, 0))
        ttk.Button(actions, text="원문 로그 보기", command=self.open_raw_log_window).pack(side="left", padx=(8, 0))
        ttk.Label(actions, textvariable=self.extension_status_var).pack(side="right")

        self.watch_log = tk.Text(self.watch_tab, height=18, wrap="word")
        self.watch_log.grid(row=1, column=0, sticky="nsew")

    def build_tools_tab(self) -> None:
        self.tools_tab.columnconfigure(1, weight=1)
        self.tools_tab.rowconfigure(5, weight=1)

        ttk.Label(
            self.tools_tab,
            text="문제가 생겼을 때 쓰는 보조 기능입니다. 일반 작업은 1~3단계만 사용하면 됩니다.",
            wraplength=820,
        ).grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 14))

        self.add_file_row(self.tools_tab, 1, "JSON 파일 1개", self.single_json_var, self.choose_single_json)

        actions = ttk.Frame(self.tools_tab)
        actions.grid(row=2, column=0, columnspan=3, sticky="ew", pady=(14, 8))
        ttk.Button(actions, text="엑셀 1개 만들기", command=self.create_excel).pack(side="left")
        ttk.Button(actions, text="다운로드 폴더 일괄 변환", command=self.create_batch_excels).pack(side="left", padx=(8, 0))

        self.tools_log = tk.Text(self.tools_tab, height=14, wrap="word")
        self.tools_log.grid(row=5, column=0, columnspan=3, sticky="nsew")

    def build_help_tab(self) -> None:
        self.help_tab.rowconfigure(0, weight=1)
        self.help_tab.columnconfigure(0, weight=1)

        text = tk.Text(self.help_tab, wrap="word")
        text.grid(row=0, column=0, sticky="nsew")
        text.insert(
            tk.END,
            "\n".join(
                [
                    f"Export Genius 자동 저장 도구 {APP_VERSION}",
                    "",
                    "기본 사용 순서:",
                    "  1. 파일 선택: 작업 목록 엑셀, 템플릿, 다운로드 폴더, 저장 폴더를 선택합니다.",
                    "  2. 작업 확인: 회사/HS코드 목록을 불러오고 저장 폴더를 만듭니다.",
                    "  3. 모니터링: 자동 실행 진행 상황을 확인합니다.",
                    "  4. GUI에서 모니터링 시작을 누르면 Edge 자동화가 실행됩니다.",
                    "  5. 새 JSON이 다운로드되면 이 프로그램이 확인 후 엑셀 파일로 저장합니다.",
                    "  6. 변환에 성공한 JSON은 결과 폴더의 _json_backup 폴더로 이동됩니다.",
                    "",
                    "Edge 확장 프로그램 확인:",
                    "  - 확장프로그램은 배포 담당자가 설치합니다.",
                    "  - 모니터링 화면 오른쪽 위에서 연결 상태를 확인합니다.",
                    "  - 확장 프로그램은 임시 JSON 다운로드를 담당하고, 이 프로그램은 엑셀 변환 후 JSON을 정리합니다.",
                    "",
                    "설정 및 로그:",
                    f"  - 설정: {SETTINGS_PATH}",
                    f"  - 로그: {LOG_DIR}",
                    "",
                    "중요한 기준:",
                    "  목표 파일 수는 방문한 회사 수가 아니라 실제 저장된 엑셀 파일 수입니다.",
                ]
            ),
        )
        text.configure(state="disabled")

    def add_file_row(self, parent: ttk.Frame, row: int, label: str, variable: tk.StringVar, command) -> None:
        ttk.Label(parent, text=label, width=16).grid(row=row, column=0, sticky="w", pady=5)
        ttk.Entry(parent, textvariable=variable).grid(row=row, column=1, sticky="ew", padx=(8, 8), pady=5)
        ttk.Button(parent, text="찾기", command=command).grid(row=row, column=2, sticky="e", pady=5)

    def choose_input(self) -> None:
        path = filedialog.askopenfilename(
            title="작업 목록 엑셀 선택",
            filetypes=[("엑셀 파일", "*.xlsx")],
            initialdir=str(DEFAULT_DOWNLOADS),
        )
        if path:
            self.input_var.set(path)
            self.persist_settings()

    def choose_template(self) -> None:
        initial = Path(self.template_var.get()).parent if self.template_var.get() else Path.home()
        path = filedialog.askopenfilename(
            title="결과 템플릿 선택",
            filetypes=[("엑셀 파일", "*.xlsx")],
            initialdir=str(initial),
        )
        if path:
            self.template_var.set(path)
            self.persist_settings()

    def choose_output_dir(self) -> None:
        path = filedialog.askdirectory(
            title="최종 저장 폴더 선택",
            initialdir=self.output_dir_var.get() or str(Path.home()),
        )
        if path:
            self.output_dir_var.set(path)
            self.persist_settings()

    def choose_single_json(self) -> None:
        path = filedialog.askopenfilename(
            title="JSON 파일 선택",
            filetypes=[("JSON 파일", "*.json")],
            initialdir=str(DEFAULT_DOWNLOADS),
        )
        if path:
            self.single_json_var.set(path)

    def validate_setup(self) -> None:
        try:
            input_path = Path(self.input_var.get())
            template_path = Path(self.template_var.get())
            download_dir = Path(self.download_dir_var.get())
            output_dir = Path(self.output_dir_var.get())

            checks = [
                ("작업 목록 엑셀", input_path.exists(), input_path),
                ("결과 템플릿", template_path.exists(), template_path),
                ("자동 확인 폴더", download_dir.exists(), download_dir),
                ("최종 저장 폴더", True, output_dir),
            ]

            lines = [
                f"프로그램 버전: {APP_VERSION}",
                f"확장 연결 주소: http://{LOCAL_SERVER_HOST}:{LOCAL_SERVER_PORT}",
                "",
            ]
            for label, ok, path in checks:
                lines.append(f"{'정상' if ok else '확인 필요'} - {label}: {path}")

            if input_path.exists():
                tasks = read_input_tasks(input_path)
                lines.append("")
                lines.append(f"작업 수: {len(tasks)}")
                lines.append(f"업체 수: {len({task.company for task in tasks})}")

            self.setup_summary.delete("1.0", tk.END)
            self.setup_summary.insert(tk.END, "\n".join(lines))
            self.persist_settings()
            self.status_var.set("선택 내용을 확인했습니다.")
        except Exception as error:
            self.status_var.set(f"오류: {error}")
            messagebox.showerror("오류", str(error))

    def go_tasks(self) -> None:
        self.preview_input()
        self.notebook.select(self.tasks_tab)

    def go_watch(self) -> None:
        self.show_selected_task()
        self.notebook.select(self.watch_tab)

    def load_tasks(self) -> list[ExportTask]:
        input_path = Path(self.input_var.get())
        if not input_path.exists():
            raise FileNotFoundError("작업 목록 엑셀 파일을 찾을 수 없습니다.")

        self.tasks = read_input_tasks(input_path)
        if not self.tasks:
            raise ValueError("작업 목록 엑셀에서 업체명/HS코드 작업을 찾지 못했습니다.")

        return self.tasks

    def rebuild_task_rows(self, tasks: list[ExportTask]) -> None:
        try:
            default_target = max(1, int(self.default_target_var.get()))
        except ValueError:
            default_target = 60
            self.default_target_var.set(str(default_target))

        self.task_rows = [
            {"selected": True, "task": task, "target": default_target, "status": "대기"}
            for task in tasks
        ]
        self.refresh_task_tree()

    def refresh_task_tree(self) -> None:
        if not hasattr(self, "tasks_tree"):
            return

        selected_items = set(self.tasks_tree.selection())
        for item in self.tasks_tree.get_children():
            self.tasks_tree.delete(item)

        for index, row in enumerate(self.task_rows, start=1):
            task = row["task"]
            item_id = f"task-{index}"
            self.tasks_tree.insert(
                "",
                "end",
                iid=item_id,
                values=(
                    "☑" if row["selected"] else "☐",
                    index,
                    f"{task.company_order:03d}. {task.company}" if task.company_order else task.company,
                    f"{task.hscode_order:03d}. {task.hscode}" if task.hscode_order else task.hscode,
                    task.min_usd,
                    task.max_usd,
                    row["target"],
                    row["status"],
                ),
            )
            if item_id in selected_items:
                self.tasks_tree.selection_add(item_id)

    def task_row_index_from_item(self, item_id: str) -> int | None:
        try:
            index = int(str(item_id).split("-", 1)[1]) - 1
        except (IndexError, ValueError):
            return None
        return index if 0 <= index < len(self.task_rows) else None

    def on_task_tree_click(self, event) -> None:
        if self.tasks_tree.identify("region", event.x, event.y) != "cell":
            return

        item_id = self.tasks_tree.identify_row(event.y)
        if self.tasks_tree.identify_column(event.x) != "#1" or not item_id:
            return

        index = self.task_row_index_from_item(item_id)
        if index is not None:
            self.task_rows[index]["selected"] = not self.task_rows[index]["selected"]
            self.refresh_task_tree()

    def select_all_tasks(self) -> None:
        for row in self.task_rows:
            row["selected"] = True
        self.refresh_task_tree()
        self.status_var.set("모든 작업을 선택했습니다.")

    def clear_all_tasks(self) -> None:
        for row in self.task_rows:
            row["selected"] = False
        self.refresh_task_tree()
        self.status_var.set("모든 작업 선택을 해제했습니다.")

    def selected_task_row_indexes(self) -> list[int]:
        selected_items = self.tasks_tree.selection() if hasattr(self, "tasks_tree") else []
        indexes = [self.task_row_index_from_item(item_id) for item_id in selected_items]
        return [index for index in indexes if index is not None]

    def apply_target_to_selected_rows(self) -> None:
        try:
            target = max(1, int(self.default_target_var.get()))
        except ValueError as error:
            raise ValueError("기본 목표 개수는 숫자로 입력해야 합니다.") from error

        indexes = [index for index, row in enumerate(self.task_rows) if row["selected"]]
        if not indexes:
            messagebox.showwarning("선택 필요", "목표 개수를 적용할 작업을 선택하세요.")
            return

        for index in indexes:
            self.task_rows[index]["target"] = target
        self.refresh_task_tree()
        self.status_var.set(f"{len(indexes)}개 작업의 목표를 {target}개로 바꿨습니다.")

    def prepare_selected_queue(self) -> None:
        try:
            if not self.task_rows:
                self.rebuild_task_rows(self.load_tasks())

            selected_indexes = [index for index, row in enumerate(self.task_rows) if row["selected"]]
            if not selected_indexes:
                raise ValueError("자동 실행할 작업을 하나 이상 체크하세요.")

            output_root = Path(self.output_dir_var.get())
            output_root.mkdir(parents=True, exist_ok=True)
            resumable_indexes = []
            for index in selected_indexes:
                task = self.task_rows[index]["task"]
                target = int(self.task_rows[index]["target"])
                task_dir = task_output_dir(output_root, task)
                task_dir.mkdir(parents=True, exist_ok=True)
                saved_count = count_task_xlsx_files(output_root, task)

                if saved_count >= target:
                    self.task_rows[index]["status"] = f"이미 완료 {saved_count}/{target}"
                else:
                    self.task_rows[index]["status"] = f"이어하기 {saved_count}/{target}"
                    resumable_indexes.append(index)

            if not resumable_indexes:
                self.refresh_task_tree()
                self.notebook.select(self.watch_tab)
                self.watch_log.delete("1.0", tk.END)
                self.append_watch_line("선택한 작업은 이미 목표 개수만큼 엑셀이 저장되어 있습니다.")
                self.status_var.set("이어할 작업 없음")
                return

            with self.queue_lock:
                self.queue_indices = resumable_indexes
                self.queue_position = 0
                self.queue_active = True
                self.queue_run_id = str(int(time.time() * 1000))
                self.queue_task_ids = {
                    row_index: f"{self.queue_run_id}:{position + 1}:{row_index + 1}"
                    for position, row_index in enumerate(self.queue_indices)
                }
                first_index = self.queue_indices[0]
                first_row = self.task_rows[first_index]
                self.set_current_task_payload(
                    self.task_payload(
                        first_row["task"],
                        first_index + 1,
                        first_row["target"],
                        self.queue_task_ids.get(first_index, ""),
                    )
                )
                self.task_rows[first_index]["status"] = "준비됨"

            self.refresh_task_tree()
            self.notebook.select(self.watch_tab)
            self.watch_log.delete("1.0", tk.END)
            self.append_watch_line(f"이어하기 준비 완료: 부족한 작업 {len(resumable_indexes)}개")
            self.append_watch_line("이미 저장된 엑셀 파일은 그대로 인정하고, 부족한 개수만 더 수집합니다.")
            self.status_var.set(f"이어하기 준비 완료: {len(resumable_indexes)}개 작업")
        except Exception as error:
            self.status_var.set(f"오류: {error}")
            messagebox.showerror("오류", str(error))

    def preview_input(self) -> None:
        try:
            tasks = self.load_tasks()
            self.rebuild_task_rows(tasks)
            if tasks:
                self.update_current_task_payload(tasks[0], 1)
            self.status_var.set(f"작업 {len(tasks)}개를 불러왔습니다.")
        except Exception as error:
            self.status_var.set(f"오류: {error}")
            messagebox.showerror("오류", str(error))

    def create_folders(self) -> None:
        try:
            tasks = self.load_tasks()
            output_root = Path(self.output_dir_var.get())
            output_root.mkdir(parents=True, exist_ok=True)

            for task in tasks:
                task_output_dir(output_root, task).mkdir(parents=True, exist_ok=True)

            self.status_var.set(f"작업 {len(tasks)}개에 대한 폴더를 만들었습니다.")
            messagebox.showinfo("완료", f"아래 위치에 회사/HS코드 폴더를 만들었습니다.\n\n{output_root}")
        except Exception as error:
            self.status_var.set(f"오류: {error}")
            messagebox.showerror("오류", str(error))

    def selected_task(self) -> ExportTask:
        tasks = self.load_tasks()

        try:
            index = int(self.task_index_var.get()) - 1
        except ValueError as error:
            raise ValueError("작업 번호는 숫자로 입력해야 합니다.") from error

        if index < 0 or index >= len(tasks):
            raise IndexError(f"작업 번호는 1부터 {len(tasks)} 사이여야 합니다.")

        return tasks[index]

    def selected_task_target_count(self) -> int:
        try:
            index = int(self.task_index_var.get()) - 1
        except ValueError:
            index = 0

        if 0 <= index < len(self.task_rows):
            return max(1, int(self.task_rows[index]["target"]))

        return max(1, int(self.default_target_var.get() or "60"))

    def show_selected_task(self) -> None:
        try:
            task = self.selected_task()
            task_number = int(self.task_index_var.get())
            self.update_current_task_payload(task, task_number)
            task_dir = task_output_dir(Path(self.output_dir_var.get()), task)
            existing = count_xlsx_files(task_dir)
            lines = [
                f"작업 번호: {self.task_index_var.get()}",
                f"업체명: {task.company}",
                f"HS코드: {task.hscode}",
                f"최소 금액: {task.min_usd}",
                f"최대 금액: {task.max_usd}",
                f"저장 위치: {task_dir}",
                f"이미 저장된 엑셀: {existing}개",
                "",
                "Edge 도우미 패널에 아래 값을 입력하세요:",
                f"  HS: {task.hscode}",
                f"  Min: {task.min_usd}",
                f"  Max: {task.max_usd}",
            ]

            self.watch_log.delete("1.0", tk.END)
            self.watch_log.insert(tk.END, "\n".join(lines) + "\n")
            self.status_var.set(f"선택 작업: {task.company} / {task.hscode}")
        except Exception as error:
            self.status_var.set(f"오류: {error}")
            messagebox.showerror("오류", str(error))

    def append_watch_line(self, line: str) -> None:
        self.watch_log.insert(tk.END, f"{line}\n")
        self.watch_log.see(tk.END)
        self.write_file_log(line)

    def write_file_log(self, line: str) -> None:
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
        try:
            with self.file_log_lock:
                LOG_DIR.mkdir(parents=True, exist_ok=True)
                rotate_log_file(APP_LOG_PATH)
                with APP_LOG_PATH.open("a", encoding="utf-8") as stream:
                    stream.write(f"[{timestamp}] {line}\n")
        except OSError:
            return

    def write_raw_log(self, entry: dict) -> None:
        try:
            with self.file_log_lock:
                LOG_DIR.mkdir(parents=True, exist_ok=True)
                rotate_log_file(RAW_LOG_PATH)
                with RAW_LOG_PATH.open("a", encoding="utf-8") as stream:
                    stream.write(json.dumps(entry, ensure_ascii=False) + "\n")
        except OSError:
            return

    def post_watch_message(self, line: str, status: str | None = None) -> None:
        self.root.after(0, lambda: self.append_watch_line(line))
        if status:
            self.root.after(0, lambda: self.status_var.set(status))

    def raw_log_text(self) -> str:
        with self.raw_log_lock:
            logs = list(self.raw_logs)

        if not logs:
            return "아직 수신된 원문 로그가 없습니다."

        return json.dumps(list(reversed(logs)), ensure_ascii=False, indent=2)

    def open_raw_log_window(self) -> None:
        window = tk.Toplevel(self.root)
        window.title("개발자용 원문 로그")
        window.geometry("920x640")
        window.minsize(720, 420)
        window.rowconfigure(1, weight=1)
        window.columnconfigure(0, weight=1)

        toolbar = ttk.Frame(window, padding=(10, 10, 10, 6))
        toolbar.grid(row=0, column=0, sticky="ew")

        text = tk.Text(window, wrap="none")
        text.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 10))

        y_scroll = ttk.Scrollbar(window, orient="vertical", command=text.yview)
        y_scroll.grid(row=1, column=1, sticky="ns", pady=(0, 10))
        x_scroll = ttk.Scrollbar(window, orient="horizontal", command=text.xview)
        x_scroll.grid(row=2, column=0, sticky="ew", padx=10)
        text.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        def refresh() -> None:
            text.delete("1.0", tk.END)
            text.insert(tk.END, self.raw_log_text())

        ttk.Button(toolbar, text="새로고침", command=refresh).pack(side="left")
        ttk.Label(toolbar, text="최근 30개 원문 로그를 표시합니다.").pack(side="left", padx=(10, 0))
        refresh()

    def queue_extension_command(self, action: str) -> dict:
        with self.command_lock:
            self.command_counter += 1
            command = {
                "id": self.command_counter,
                "action": action,
                "createdAt": time.time(),
                "claimed": False,
            }
            self.pending_command = command
            return dict(command)

    def check_extension_command_received(self, command_id: int) -> None:
        with self.command_lock:
            pending = self.pending_command
            waiting = bool(
                pending
                and pending.get("id") == command_id
                and not pending.get("claimed")
            )

        if waiting:
            message = (
                "확장프로그램이 시작 명령을 받지 못했습니다. "
                "Edge에서 Export Genius 검색 결과 페이지가 열려 있는지 확인하세요."
            )
            self.append_watch_line(message)
            self.status_var.set("확장프로그램 연결 확인 필요")
            self.extension_status_var.set("확장프로그램 응답 없음")

    def start_gui_automation(self) -> None:
        try:
            if not self.local_server:
                raise RuntimeError("로컬 연결 서버가 실행되지 않아 자동화를 시작할 수 없습니다.")
            if self.last_extension_version and self.last_extension_version != EXPECTED_EXTENSION_VERSION:
                raise RuntimeError(
                    f"확장프로그램 버전이 맞지 않습니다. "
                    f"필요 버전: {EXPECTED_EXTENSION_VERSION}, 연결 버전: {self.last_extension_version}"
                )
            if not self.queue_active:
                self.prepare_selected_queue()

            self.stop_requested = False
            command = self.queue_extension_command("startQueue")
            self.append_watch_line("")
            self.append_watch_line("브라우저 자동 실행을 요청했습니다. Export Genius 페이지가 열려 있으면 곧 작업이 시작됩니다.")
            self.status_var.set("브라우저 자동 실행 대기 중")
            self.root.after(10000, lambda command_id=command["id"]: self.check_extension_command_received(command_id))
        except Exception as error:
            self.status_var.set(f"오류: {error}")
            messagebox.showerror("오류", str(error))

    def stop_gui_automation(self) -> None:
        self.stop_requested = True
        self.queue_extension_command("stop")
        self.status_var.set("중단 요청을 보냈습니다.")
        self.append_watch_line("중단 요청을 보냈습니다. 현재 처리 단계가 멈출 수 있는 지점에 도달하면 작업이 중단됩니다.")

    def start_watch_task(self) -> None:
        if self.watch_thread and self.watch_thread.is_alive():
            messagebox.showwarning("모니터링 중", "이미 모니터링 중인 작업이 있습니다.")
            return

        try:
            task = self.selected_task()
            template_path = Path(self.template_var.get())
            json_dir = Path(self.download_dir_var.get())
            output_root = Path(self.output_dir_var.get())
            target_count = self.selected_task_target_count()
            next_file_timeout_seconds = 300

            if not template_path.exists():
                raise FileNotFoundError("결과 템플릿 파일을 찾을 수 없습니다.")
            if not json_dir.exists():
                raise FileNotFoundError("다운로드 폴더를 찾을 수 없습니다.")
            if target_count <= 0:
                raise ValueError("목표 파일 수는 1 이상이어야 합니다.")
            task_dir = task_output_dir(output_root, task)
            task_dir.mkdir(parents=True, exist_ok=True)

            self.stop_event.clear()
            self.show_selected_task()
            self.append_watch_line("")
            self.append_watch_line(f"다운로드 모니터링을 시작합니다. 목표: 엑셀 {target_count}개")
            self.append_watch_line(f"다음 엑셀 파일이 {next_file_timeout_seconds}초 안에 생성되지 않으면 중단합니다.")

            self.watch_thread = threading.Thread(
                target=self.watch_task_worker,
                args=(task, template_path, json_dir, task_dir, target_count, next_file_timeout_seconds),
                daemon=True,
            )
            self.watch_thread.start()
            self.status_var.set("다운로드 폴더를 모니터링 중입니다.")
        except Exception as error:
            self.status_var.set(f"오류: {error}")
            messagebox.showerror("오류", str(error))

    def stop_watch_task(self) -> None:
        self.stop_event.set()
        self.status_var.set("모니터링 작업을 중단하는 중입니다.")

    def watch_task_worker(
        self,
        task: ExportTask,
        template_path: Path,
        json_dir: Path,
        task_dir: Path,
        target_count: int,
        next_file_timeout_seconds: int,
    ) -> None:
        started_at = time.time()
        last_progress_at = started_at
        seen_paths = {path.resolve() for path in json_dir.glob("*.json")}

        try:
            while not self.stop_event.is_set():
                current_count = count_xlsx_files(task_dir)
                if current_count >= target_count:
                    self.post_watch_message(f"완료: 엑셀 {current_count}/{target_count}개 저장됨", "모니터링 완료")
                    return

                if time.time() - last_progress_at > next_file_timeout_seconds:
                    raise TimeoutError(
                        f"다음 엑셀 파일이 {next_file_timeout_seconds}초 동안 생성되지 않았습니다. "
                        f"현재 엑셀 {current_count}/{target_count}개가 저장되었습니다."
                    )

                new_files = sorted(
                    [
                        path
                        for path in json_dir.glob("*.json")
                        if path.resolve() not in seen_paths and path.stat().st_mtime >= started_at
                    ],
                    key=lambda path: path.stat().st_mtime,
                )

                if not new_files:
                    time.sleep(1)
                    continue

                for json_path in new_files:
                    seen_paths.add(json_path.resolve())
                    try:
                        self.process_downloaded_json(task, template_path, task_dir, json_path)
                    except Exception:
                        if json_path.exists():
                            failed_path = move_json_to_folder(json_path, task_dir / "_failed_json")
                            self.post_watch_message(f"실패한 JSON 보관: {failed_path}")
                        raise

                    current_count = count_xlsx_files(task_dir)
                    last_progress_at = time.time()
                    self.post_watch_message(f"저장 완료 {current_count}/{target_count}: {json_path.name}", f"저장 {current_count}/{target_count}")

                    if current_count >= target_count:
                        self.post_watch_message(f"완료: 엑셀 {current_count}/{target_count}개 저장됨", "모니터링 완료")
                        return

            self.post_watch_message("사용자가 모니터링을 중단했습니다.", "모니터링 중단")
        except Exception as error:
            error_text = str(error)
            self.post_watch_message(f"오류: {error_text}", f"오류: {error_text}")
            self.root.after(0, lambda: messagebox.showerror("모니터링 오류", error_text))

    def process_downloaded_json(self, task: ExportTask, template_path: Path, task_dir: Path, json_path: Path) -> Path:
        data = wait_until_json_ready(json_path)
        json_hscode = normalize_hscode(data.get("hsCode"))

        if data.get("qualified") is False:
            raise ValueError(f"{json_path.name} 파일은 조건을 통과하지 않은 JSON입니다.")
        if json_hscode and json_hscode != task.hscode:
            raise ValueError(f"{json_path.name} 파일의 HS코드는 {json_hscode}입니다. 현재 작업 HS코드 {task.hscode}와 다릅니다.")

        buyer_name = data.get("excel", {}).get("pink", {}).get("Company_Name") or json_path.stem
        if not json_filename_matches_buyer(json_path, buyer_name):
            raise ValueError(f"{json_path.name} 파일명과 JSON 내부 바이어명({buyer_name})이 일치하지 않습니다.")

        output_path = next_output_path(
            task_dir,
            buyer_name,
            task_output_dirs(Path(self.output_dir_var.get()), task),
        )
        fill_workbook(template_path, json_path, output_path)

        move_json_to_folder(json_path, task_dir / "_json_backup")
        return output_path

    def create_excel(self) -> None:
        try:
            template_path = Path(self.template_var.get())
            json_path = Path(self.single_json_var.get())
            output_dir = Path(self.output_dir_var.get())

            if not template_path.exists():
                raise FileNotFoundError("결과 템플릿 파일을 찾을 수 없습니다.")
            if not json_path.exists():
                raise FileNotFoundError("JSON 파일을 찾을 수 없습니다.")

            output_dir.mkdir(parents=True, exist_ok=True)
            data = load_data(json_path)
            company_name = data.get("excel", {}).get("pink", {}).get("Company_Name") or json_path.stem
            output_path = output_dir / f"{safe_filename(company_name)}.xlsx"

            fill_workbook(template_path, json_path, output_path)
            self.tools_log.insert(tk.END, f"생성 완료: {output_path}\n")
            self.status_var.set(f"생성 완료: {output_path}")
            messagebox.showinfo("완료", f"엑셀 파일을 만들었습니다.\n\n{output_path}")
        except Exception as error:
            self.status_var.set(f"오류: {error}")
            messagebox.showerror("오류", str(error))

    def create_batch_excels(self) -> None:
        try:
            template_path = Path(self.template_var.get())
            json_dir = Path(self.download_dir_var.get())
            output_dir = Path(self.output_dir_var.get())

            if not template_path.exists():
                raise FileNotFoundError("결과 템플릿 파일을 찾을 수 없습니다.")
            if not json_dir.exists():
                raise FileNotFoundError("다운로드 폴더를 찾을 수 없습니다.")

            output_dir.mkdir(parents=True, exist_ok=True)
            json_files = sorted(json_dir.glob("*.json"), key=lambda path: path.stat().st_mtime)
            if not json_files:
                raise FileNotFoundError("다운로드 폴더에서 JSON 파일을 찾지 못했습니다.")

            created = 0
            skipped = 0
            errors: list[str] = []

            for json_path in json_files:
                try:
                    data = load_data(json_path)
                    if data.get("qualified") is False:
                        skipped += 1
                        continue

                    company_name = data.get("excel", {}).get("pink", {}).get("Company_Name") or json_path.stem
                    output_path = output_dir / f"{created + 1}. {safe_filename(company_name)}.xlsx"
                    fill_workbook(template_path, json_path, output_path)
                    created += 1
                    self.tools_log.insert(tk.END, f"생성 완료: {output_path}\n")
                except Exception as error:
                    errors.append(f"{json_path.name}: {error}")

            message = f"엑셀 {created}개 생성, JSON {skipped}개 제외"
            if errors:
                message += f", 오류 {len(errors)}개"
                self.tools_log.insert(tk.END, "\n".join(errors) + "\n")

            self.status_var.set(message)
            messagebox.showinfo("일괄 변환 완료", message)
        except Exception as error:
            self.status_var.set(f"오류: {error}")
            messagebox.showerror("오류", str(error))


def main() -> None:
    root = tk.Tk()
    ExcelGui(root)
    root.mainloop()


if __name__ == "__main__":
    main()
