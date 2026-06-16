import argparse
import json
import re
import shutil
from pathlib import Path

from openpyxl import load_workbook


PINK_LABELS = {
    "Company_Name": "Company_Name",
    "Annual_Import_Turnover": "Annual_Import_Turnover",
    "Annual_Import_Shipment": "Annual_Import_Shipment",
    "Annual_Export_Turnover": "Annual_Export_Turnover",
    "Country_of_Origin": "Country_of_Origin",
    "Address": "Address",
    "Homepage": "Homepage",
    "Phone_Number": "Phone_Number",
}


def safe_filename(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1F]', " ", value or "company")
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned[:120] or "company"


def load_data(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as file:
        return json.load(file)


def set_cell(ws, row: int, col: int, value):
    ws.cell(row=row, column=col).value = value if value not in (None, "") else None


def find_rows(ws, label: str) -> list[int]:
    return [
        row
        for row in range(1, ws.max_row + 1)
        if str(ws.cell(row=row, column=1).value or "").strip() == label
    ]


def fill_repeated_rows(ws, label: str, rows_data: list[dict], limit: int):
    rows = find_rows(ws, label)[:limit]

    for index, row in enumerate(rows):
        item = rows_data[index] if index < len(rows_data) else {}
        set_cell(ws, row, 2, item.get("Value1"))
        set_cell(ws, row, 3, item.get("Value2"))


def fill_workbook(template_path: Path, data_path: Path, output_path: Path):
    data = load_data(data_path)
    excel = data.get("excel", {})
    pink = excel.get("pink", {})
    blue = excel.get("blue", {})
    commodity = excel.get("commodity", {})

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)

    wb = load_workbook(output_path)
    ws = wb.active

    for row in range(1, ws.max_row + 1):
        label = str(ws.cell(row=row, column=1).value or "").strip()
        key = PINK_LABELS.get(label)
        if key:
            set_cell(ws, row, 2, pink.get(key))
            set_cell(ws, row, 3, None)

    fill_repeated_rows(ws, "Supplier_Country", blue.get("Supplier_Country", []), 5)
    fill_repeated_rows(ws, "Buyer_Country", blue.get("Buyer_Country", []), 5)
    fill_repeated_rows(ws, "Import", commodity.get("Import", []), 10)
    fill_repeated_rows(ws, "Export", commodity.get("Export", []), 10)

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="Fill the fixed Export Genius Excel template without changing its layout."
    )
    parser.add_argument("--template", required=True, help="Path to the existing xlsx template.")
    parser.add_argument("--json", required=True, help="Path to the JSON downloaded by the extension.")
    parser.add_argument("--output", help="Output xlsx path. Defaults to '<Company_Name>.xlsx' next to the JSON.")
    args = parser.parse_args()

    template_path = Path(args.template)
    data_path = Path(args.json)
    data = load_data(data_path)
    company_name = data.get("excel", {}).get("pink", {}).get("Company_Name") or data_path.stem
    output_path = Path(args.output) if args.output else data_path.with_name(f"{safe_filename(company_name)}.xlsx")

    fill_workbook(template_path, data_path, output_path)
    print(output_path)


if __name__ == "__main__":
    main()
